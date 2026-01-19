from flask import Flask, render_template, request, send_file, jsonify, redirect, send_from_directory
import openpyxl
import io
import os
from werkzeug.utils import secure_filename
from urllib.parse import unquote

app = Flask(__name__)
workbooks = {}

@app.route("/")
def root():
    return redirect("/app3/")

@app.route("/app3/")
def index():
    return render_template("index.html")

@app.route("/app3/sw.js")
def serve_sw():
    return send_from_directory('static/app3', 'sw.js')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    try:
        file_bytes = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        file_id = secure_filename(file.filename)
        workbooks[file_id] = {
            'filename': file.filename,
            'file_bytes': file_bytes,
            'sheet_names': wb.sheetnames
        }
        return jsonify({'success': True, 'file_id': file_id, 'sheet_names': wb.sheetnames})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_sheet_data/<file_id>/<path:sheet_name>')
def get_sheet_data(file_id, sheet_name):
    if file_id not in workbooks:
        return jsonify({'error': 'File not found'}), 404
    try:
        decoded_name = unquote(sheet_name)
        wb = openpyxl.load_workbook(io.BytesIO(workbooks[file_id]['file_bytes']), data_only=True)
        ws = wb[decoded_name]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(c) if c is not None else '' for c in row])

        headers = list(data[7]) if len(data) > 7 else []
        body = data[8:]
        headers.insert(0, 'Saisie CSV')
        
        clean_data = []
        for r in body:
            if any(str(x).strip() for x in r):
                row_list = list(r)
                row_list.insert(0, '') 
                clean_data.append(row_list)
        
        return jsonify({'headers': headers, 'data': clean_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update_sheet_in_memory', methods=['POST'])
def update_sheet_in_memory():
    req_data = request.json
    file_id = req_data.get('file_id')
    sheet_name = req_data.get('sheet_name')
    updates = req_data.get('updated_data')
    
    if file_id not in workbooks:
        return jsonify({'error': 'Session expired'}), 404

    try:
        wb = openpyxl.load_workbook(io.BytesIO(workbooks[file_id]['file_bytes']))
        ws = wb[sheet_name]
        for u in updates:
            rIdx, cIdx, val = int(u['row_index']), int(u['column_index']), u['value']
            if cIdx == 0: continue 
            ws.cell(row=9 + rIdx, column=cIdx).value = val

        output = io.BytesIO()
        wb.save(output)
        workbooks[file_id]['file_bytes'] = output.getvalue()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/save_file', methods=['POST'])
def save_file():
    req_data = request.json
    file_id = req_data.get('file_id')
    if file_id not in workbooks:
        return jsonify({'error': 'File not found'}), 404
    try:
        output = io.BytesIO(workbooks[file_id]['file_bytes'])
        return send_file(output, download_name="notes_complet.xlsx", as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8003, debug=True)
